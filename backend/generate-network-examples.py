#!/usr/bin/env python3
"""
Generator for CIM Network Excel Examples with Load Flow Data
Creates realistic power network examples for testing
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import os

def create_simple_network():
    """
    Simple 3-bus network with generator, load, and transmission line
    Good for basic load flow testing
    """
    
    # Substations
    substations = pd.DataFrame([
        {'id': 'SUB_CENTRAL', 'name': 'Central Substation', 'region': 'Central', 'voltage': 220000},
        {'id': 'SUB_NORTH', 'name': 'North Substation', 'region': 'North', 'voltage': 220000},
        {'id': 'SUB_SOUTH', 'name': 'South Substation', 'region': 'South', 'voltage': 220000},
    ])
    
    # Buses (TopologicalNodes)
    buses = pd.DataFrame([
        {'id': 'BUS_CENTRAL_220', 'name': 'Central 220kV Bus', 'substation': 'SUB_CENTRAL', 'baseVoltage': 220000, 'type': 'SLACK'},
        {'id': 'BUS_NORTH_220', 'name': 'North 220kV Bus', 'substation': 'SUB_NORTH', 'baseVoltage': 220000, 'type': 'PQ'},
        {'id': 'BUS_SOUTH_220', 'name': 'South 220kV Bus', 'substation': 'SUB_SOUTH', 'baseVoltage': 220000, 'type': 'PQ'},
    ])
    
    # Transmission Lines (with realistic electrical parameters)
    lines = pd.DataFrame([
        {'id': 'LINE_CN', 'name': 'Central-North Line', 'from': 'BUS_CENTRAL_220', 'to': 'BUS_NORTH_220', 
         'r': 0.05, 'x': 0.15, 'b': 0.002, 'length': 50, 'ratedCurrent': 1500},
        {'id': 'LINE_CS', 'name': 'Central-South Line', 'from': 'BUS_CENTRAL_220', 'to': 'BUS_SOUTH_220', 
         'r': 0.08, 'x': 0.20, 'b': 0.0025, 'length': 75, 'ratedCurrent': 1200},
        {'id': 'LINE_NS', 'name': 'North-South Line', 'from': 'BUS_NORTH_220', 'to': 'BUS_SOUTH_220', 
         'r': 0.06, 'x': 0.18, 'b': 0.003, 'length': 60, 'ratedCurrent': 1300},
    ])
    
    # Generators
    generators = pd.DataFrame([
        {'id': 'GEN_CENTRAL', 'name': 'Central Generator', 'bus': 'BUS_CENTRAL_220', 
         'pMin': 0, 'pMax': 500, 'qMin': -200, 'qMax': 200, 'p': 300, 'v': 1.02},
        {'id': 'GEN_NORTH', 'name': 'North Wind Farm', 'bus': 'BUS_NORTH_220', 
         'pMin': 0, 'pMax': 200, 'qMin': -100, 'qMax': 100, 'p': 150, 'v': 1.00},
    ])
    
    # Loads
    loads = pd.DataFrame([
        {'id': 'LOAD_NORTH', 'name': 'North City Load', 'bus': 'BUS_NORTH_220', 'p': 200, 'q': 50},
        {'id': 'LOAD_SOUTH', 'name': 'South Industrial Load', 'bus': 'BUS_SOUTH_220', 'p': 250, 'q': 80},
    ])
    
    # Voltage Levels
    voltage_levels = pd.DataFrame([
        {'id': 'VL_220_CENTRAL', 'name': '220kV Central', 'substation': 'SUB_CENTRAL', 'nominalVoltage': 220000, 'highVoltageLimit': 242000, 'lowVoltageLimit': 198000},
        {'id': 'VL_220_NORTH', 'name': '220kV North', 'substation': 'SUB_NORTH', 'nominalVoltage': 220000, 'highVoltageLimit': 242000, 'lowVoltageLimit': 198000},
        {'id': 'VL_220_SOUTH', 'name': '220kV South', 'substation': 'SUB_SOUTH', 'nominalVoltage': 220000, 'highVoltageLimit': 242000, 'lowVoltageLimit': 198000},
    ])
    
    return {
        'Substations': substations,
        'Buses': buses,
        'Lines': lines,
        'Generators': generators,
        'Loads': loads,
        'Voltage Levels': voltage_levels
    }

def create_complex_network():
    """
    Complex 6-bus network with transformers, multiple generators and loads
    For advanced load flow analysis
    """
    
    # Substations
    substations = pd.DataFrame([
        {'id': 'SUB_MAIN', 'name': 'Main Power Station', 'region': 'Central', 'voltage': 400000},
        {'id': 'SUB_EAST', 'name': 'East Substation', 'region': 'East', 'voltage': 220000},
        {'id': 'SUB_WEST', 'name': 'West Substation', 'region': 'West', 'voltage': 220000},
        {'id': 'SUB_NORTH', 'name': 'North Distribution', 'region': 'North', 'voltage': 110000},
        {'id': 'SUB_SOUTH', 'name': 'South Distribution', 'region': 'South', 'voltage': 110000},
    ])
    
    # Buses
    buses = pd.DataFrame([
        {'id': 'BUS_MAIN_400', 'name': 'Main 400kV Bus', 'substation': 'SUB_MAIN', 'baseVoltage': 400000, 'type': 'SLACK'},
        {'id': 'BUS_EAST_220', 'name': 'East 220kV Bus', 'substation': 'SUB_EAST', 'baseVoltage': 220000, 'type': 'PV'},
        {'id': 'BUS_WEST_220', 'name': 'West 220kV Bus', 'substation': 'SUB_WEST', 'baseVoltage': 220000, 'type': 'PQ'},
        {'id': 'BUS_NORTH_110', 'name': 'North 110kV Bus', 'substation': 'SUB_NORTH', 'baseVoltage': 110000, 'type': 'PQ'},
        {'id': 'BUS_SOUTH_110', 'name': 'South 110kV Bus', 'substation': 'SUB_SOUTH', 'baseVoltage': 110000, 'type': 'PQ'},
        {'id': 'BUS_EAST_110', 'name': 'East 110kV Bus', 'substation': 'SUB_EAST', 'baseVoltage': 110000, 'type': 'PQ'},
    ])
    
    # Transmission Lines
    lines = pd.DataFrame([
        {'id': 'LINE_MAIN_EAST', 'name': '400kV Main-East', 'from': 'BUS_MAIN_400', 'to': 'BUS_EAST_220', 
         'r': 0.02, 'x': 0.10, 'b': 0.001, 'length': 100, 'ratedCurrent': 2500},
        {'id': 'LINE_MAIN_WEST', 'name': '400kV Main-West', 'from': 'BUS_MAIN_400', 'to': 'BUS_WEST_220', 
         'r': 0.03, 'x': 0.12, 'b': 0.0015, 'length': 120, 'ratedCurrent': 2500},
        {'id': 'LINE_EAST_WEST', 'name': '220kV East-West', 'from': 'BUS_EAST_220', 'to': 'BUS_WEST_220', 
         'r': 0.05, 'x': 0.15, 'b': 0.002, 'length': 80, 'ratedCurrent': 1500},
        {'id': 'LINE_NORTH_SOUTH', 'name': '110kV North-South', 'from': 'BUS_NORTH_110', 'to': 'BUS_SOUTH_110', 
         'r': 0.10, 'x': 0.25, 'b': 0.003, 'length': 50, 'ratedCurrent': 800},
    ])
    
    # Transformers
    transformers = pd.DataFrame([
        {'id': 'XFMR_EAST', 'name': 'East 220/110kV Transformer', 'from': 'BUS_EAST_220', 'to': 'BUS_EAST_110',
         'r': 0.005, 'x': 0.08, 'ratedS': 300, 'ratio': 2.0},
        {'id': 'XFMR_NORTH', 'name': 'North 220/110kV Transformer', 'from': 'BUS_EAST_220', 'to': 'BUS_NORTH_110',
         'r': 0.005, 'x': 0.08, 'ratedS': 200, 'ratio': 2.0},
        {'id': 'XFMR_SOUTH', 'name': 'South 220/110kV Transformer', 'from': 'BUS_WEST_220', 'to': 'BUS_SOUTH_110',
         'r': 0.005, 'x': 0.08, 'ratedS': 200, 'ratio': 2.0},
    ])
    
    # Generators
    generators = pd.DataFrame([
        {'id': 'GEN_MAIN_1', 'name': 'Main Generator Unit 1', 'bus': 'BUS_MAIN_400', 
         'pMin': 100, 'pMax': 1000, 'qMin': -400, 'qMax': 400, 'p': 800, 'v': 1.03},
        {'id': 'GEN_MAIN_2', 'name': 'Main Generator Unit 2', 'bus': 'BUS_MAIN_400', 
         'pMin': 100, 'pMax': 1000, 'qMin': -400, 'qMax': 400, 'p': 700, 'v': 1.03},
        {'id': 'GEN_EAST_WIND', 'name': 'East Wind Farm', 'bus': 'BUS_EAST_220', 
         'pMin': 0, 'pMax': 300, 'qMin': -150, 'qMax': 150, 'p': 250, 'v': 1.00},
        {'id': 'GEN_WEST_SOLAR', 'name': 'West Solar Park', 'bus': 'BUS_WEST_220', 
         'pMin': 0, 'pMax': 200, 'qMin': -100, 'qMax': 100, 'p': 180, 'v': 1.00},
    ])
    
    # Loads
    loads = pd.DataFrame([
        {'id': 'LOAD_EAST', 'name': 'East City Load', 'bus': 'BUS_EAST_110', 'p': 350, 'q': 120},
        {'id': 'LOAD_WEST', 'name': 'West City Load', 'bus': 'BUS_WEST_220', 'p': 400, 'q': 150},
        {'id': 'LOAD_NORTH', 'name': 'North Residential', 'bus': 'BUS_NORTH_110', 'p': 250, 'q': 80},
        {'id': 'LOAD_SOUTH', 'name': 'South Industrial', 'bus': 'BUS_SOUTH_110', 'p': 450, 'q': 180},
        {'id': 'LOAD_MAIN', 'name': 'Main Auxiliary Load', 'bus': 'BUS_MAIN_400', 'p': 50, 'q': 20},
    ])
    
    # Voltage Levels
    voltage_levels = pd.DataFrame([
        {'id': 'VL_400_MAIN', 'name': '400kV Main', 'substation': 'SUB_MAIN', 'nominalVoltage': 400000, 'highVoltageLimit': 420000, 'lowVoltageLimit': 380000},
        {'id': 'VL_220_EAST', 'name': '220kV East', 'substation': 'SUB_EAST', 'nominalVoltage': 220000, 'highVoltageLimit': 242000, 'lowVoltageLimit': 198000},
        {'id': 'VL_220_WEST', 'name': '220kV West', 'substation': 'SUB_WEST', 'nominalVoltage': 220000, 'highVoltageLimit': 242000, 'lowVoltageLimit': 198000},
        {'id': 'VL_110_NORTH', 'name': '110kV North', 'substation': 'SUB_NORTH', 'nominalVoltage': 110000, 'highVoltageLimit': 121000, 'lowVoltageLimit': 99000},
        {'id': 'VL_110_SOUTH', 'name': '110kV South', 'substation': 'SUB_SOUTH', 'nominalVoltage': 110000, 'highVoltageLimit': 121000, 'lowVoltageLimit': 99000},
        {'id': 'VL_110_EAST', 'name': '110kV East', 'substation': 'SUB_EAST', 'nominalVoltage': 110000, 'highVoltageLimit': 121000, 'lowVoltageLimit': 99000},
    ])
    
    return {
        'Substations': substations,
        'Buses': buses,
        'Lines': lines,
        'Transformers': transformers,
        'Generators': generators,
        'Loads': loads,
        'Voltage Levels': voltage_levels
    }

def style_excel_sheet(ws, df):
    """Apply styling to Excel worksheet"""
    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

def save_to_excel(data, filename):
    """Save network data to Excel file with styling"""
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        for sheet_name, df in data.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Apply styling
            ws = writer.sheets[sheet_name]
            style_excel_sheet(ws, df)
    
    print(f"✅ Created: {filename}")
    print(f"   Sheets: {', '.join(data.keys())}")
    print(f"   Total entities: {sum(len(df) for df in data.values())}")
    print()

def main():
    """Generate example network files"""
    print("🔌 Generating CIM Network Examples for Load Flow Analysis")
    print("=" * 60)
    print()
    
    # Create output directory
    output_dir = '/Users/zakaria/Documents/CIM-SemanticGraph-Platform/backend/examples'
    os.makedirs(output_dir, exist_ok=True)
    
    # Generate simple network
    print("📊 Generating Simple Network (3-bus)...")
    simple_network = create_simple_network()
    save_to_excel(simple_network, f'{output_dir}/network-simple-3bus.xlsx')
    
    # Generate complex network
    print("📊 Generating Complex Network (6-bus)...")
    complex_network = create_complex_network()
    save_to_excel(complex_network, f'{output_dir}/network-complex-6bus.xlsx')
    
    print("=" * 60)
    print("✅ All network examples generated successfully!")
    print()
    print("📁 Files created in:", output_dir)
    print()
    print("🧪 Test these files with:")
    print("   curl -X POST http://localhost:8080/api/excel/import \\")
    print("     -u admin:admin123 \\")
    print("     -F 'file=@examples/network-simple-3bus.xlsx'")
    print()

if __name__ == '__main__':
    main()
