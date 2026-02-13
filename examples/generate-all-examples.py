#!/usr/bin/env python3
"""
Générateur complet de fichiers d'exemple pour tests Load Flow
Crée des fichiers Excel, RDF et XML de différentes tailles
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os
import xml.etree.ElementTree as ET
from xml.dom import minidom

def create_small_network():
    """Réseau petit : 2 buses - Test minimal"""
    substations = pd.DataFrame([
        {'id': 'SUB_A', 'name': 'Substation A', 'region': 'Region A', 'voltage': 220000},
        {'id': 'SUB_B', 'name': 'Substation B', 'region': 'Region B', 'voltage': 220000},
    ])
    
    buses = pd.DataFrame([
        {'id': 'BUS_A_220', 'name': 'Bus A 220kV', 'substation': 'SUB_A', 'baseVoltage': 220000, 'type': 'SLACK'},
        {'id': 'BUS_B_220', 'name': 'Bus B 220kV', 'substation': 'SUB_B', 'baseVoltage': 220000, 'type': 'PQ'},
    ])
    
    lines = pd.DataFrame([
        {'id': 'LINE_AB', 'name': 'Line A-B', 'from': 'BUS_A_220', 'to': 'BUS_B_220', 
         'r': 0.05, 'x': 0.15, 'b': 0.002, 'length': 50, 'ratedCurrent': 1500},
    ])
    
    generators = pd.DataFrame([
        {'id': 'GEN_A', 'name': 'Generator A', 'bus': 'BUS_A_220', 
         'pMin': 0, 'pMax': 500, 'qMin': -200, 'qMax': 200, 'p': 300, 'v': 1.02},
    ])
    
    # Charge = 280 MW (300 MW génération - 20 MW pertes estimées)
    loads = pd.DataFrame([
        {'id': 'LOAD_B', 'name': 'Load B', 'bus': 'BUS_B_220', 'p': 280, 'q': 80},
    ])
    
    voltage_levels = pd.DataFrame([
        {'id': 'VL_220_A', 'name': '220kV A', 'substation': 'SUB_A', 'nominalVoltage': 220000, 'highVoltageLimit': 242000, 'lowVoltageLimit': 198000},
        {'id': 'VL_220_B', 'name': '220kV B', 'substation': 'SUB_B', 'nominalVoltage': 220000, 'highVoltageLimit': 242000, 'lowVoltageLimit': 198000},
    ])
    
    return {
        'Substations': substations,
        'Buses': buses,
        'Lines': lines,
        'Generators': generators,
        'Loads': loads,
        'Voltage Levels': voltage_levels
    }

def create_medium_network():
    """Réseau moyen : 8 buses - Test réaliste"""
    substations = pd.DataFrame([
        {'id': 'SUB_CENTRAL', 'name': 'Central Station', 'region': 'Central', 'voltage': 400000},
        {'id': 'SUB_NORTH', 'name': 'North Substation', 'region': 'North', 'voltage': 220000},
        {'id': 'SUB_SOUTH', 'name': 'South Substation', 'region': 'South', 'voltage': 220000},
        {'id': 'SUB_EAST', 'name': 'East Distribution', 'region': 'East', 'voltage': 110000},
        {'id': 'SUB_WEST', 'name': 'West Distribution', 'region': 'West', 'voltage': 110000},
    ])
    
    buses = pd.DataFrame([
        {'id': 'BUS_CENTRAL_400', 'name': 'Central 400kV', 'substation': 'SUB_CENTRAL', 'baseVoltage': 400000, 'type': 'SLACK'},
        {'id': 'BUS_NORTH_220', 'name': 'North 220kV', 'substation': 'SUB_NORTH', 'baseVoltage': 220000, 'type': 'PV'},
        {'id': 'BUS_SOUTH_220', 'name': 'South 220kV', 'substation': 'SUB_SOUTH', 'baseVoltage': 220000, 'type': 'PV'},
        {'id': 'BUS_EAST_110', 'name': 'East 110kV', 'substation': 'SUB_EAST', 'baseVoltage': 110000, 'type': 'PQ'},
        {'id': 'BUS_WEST_110', 'name': 'West 110kV', 'substation': 'SUB_WEST', 'baseVoltage': 110000, 'type': 'PQ'},
        {'id': 'BUS_NORTH_110', 'name': 'North 110kV', 'substation': 'SUB_NORTH', 'baseVoltage': 110000, 'type': 'PQ'},
        {'id': 'BUS_SOUTH_110', 'name': 'South 110kV', 'substation': 'SUB_SOUTH', 'baseVoltage': 110000, 'type': 'PQ'},
        {'id': 'BUS_CENTRAL_220', 'name': 'Central 220kV', 'substation': 'SUB_CENTRAL', 'baseVoltage': 220000, 'type': 'PQ'},
    ])
    
    lines = pd.DataFrame([
        {'id': 'LINE_CENTRAL_NORTH', 'name': 'Central-North 400kV', 'from': 'BUS_CENTRAL_400', 'to': 'BUS_NORTH_220', 
         'r': 0.02, 'x': 0.10, 'b': 0.001, 'length': 100, 'ratedCurrent': 2500},
        {'id': 'LINE_CENTRAL_SOUTH', 'name': 'Central-South 400kV', 'from': 'BUS_CENTRAL_400', 'to': 'BUS_SOUTH_220', 
         'r': 0.025, 'x': 0.12, 'b': 0.0012, 'length': 120, 'ratedCurrent': 2500},
        {'id': 'LINE_NORTH_SOUTH', 'name': 'North-South 220kV', 'from': 'BUS_NORTH_220', 'to': 'BUS_SOUTH_220', 
         'r': 0.05, 'x': 0.15, 'b': 0.002, 'length': 80, 'ratedCurrent': 1500},
        {'id': 'LINE_EAST_WEST', 'name': 'East-West 110kV', 'from': 'BUS_EAST_110', 'to': 'BUS_WEST_110', 
         'r': 0.10, 'x': 0.25, 'b': 0.003, 'length': 50, 'ratedCurrent': 800},
    ])
    
    transformers = pd.DataFrame([
        {'id': 'XFMR_NORTH', 'name': 'North 220/110kV', 'from': 'BUS_NORTH_220', 'to': 'BUS_NORTH_110',
         'r': 0.005, 'x': 0.08, 'ratedS': 200, 'ratio': 2.0},
        {'id': 'XFMR_SOUTH', 'name': 'South 220/110kV', 'from': 'BUS_SOUTH_220', 'to': 'BUS_SOUTH_110',
         'r': 0.005, 'x': 0.08, 'ratedS': 200, 'ratio': 2.0},
        {'id': 'XFMR_CENTRAL', 'name': 'Central 400/220kV', 'from': 'BUS_CENTRAL_400', 'to': 'BUS_CENTRAL_220',
         'r': 0.003, 'x': 0.06, 'ratedS': 500, 'ratio': 1.82},
    ])
    
    generators = pd.DataFrame([
        {'id': 'GEN_CENTRAL_1', 'name': 'Central Gen 1', 'bus': 'BUS_CENTRAL_400', 
         'pMin': 200, 'pMax': 1000, 'qMin': -400, 'qMax': 400, 'p': 800, 'v': 1.03},
        {'id': 'GEN_CENTRAL_2', 'name': 'Central Gen 2', 'bus': 'BUS_CENTRAL_400', 
         'pMin': 200, 'pMax': 1000, 'qMin': -400, 'qMax': 400, 'p': 700, 'v': 1.03},
        {'id': 'GEN_NORTH', 'name': 'North Wind', 'bus': 'BUS_NORTH_220', 
         'pMin': 0, 'pMax': 300, 'qMin': -150, 'qMax': 150, 'p': 250, 'v': 1.00},
        {'id': 'GEN_SOUTH', 'name': 'South Solar', 'bus': 'BUS_SOUTH_220', 
         'pMin': 0, 'pMax': 200, 'qMin': -100, 'qMax': 100, 'p': 180, 'v': 1.00},
    ])
    # Total génération: 1930 MW
    
    # Charges: 1650 MW (1930 - 30 MW pertes estimées)
    loads = pd.DataFrame([
        {'id': 'LOAD_EAST', 'name': 'East City', 'bus': 'BUS_EAST_110', 'p': 350, 'q': 120},
        {'id': 'LOAD_WEST', 'name': 'West City', 'bus': 'BUS_WEST_110', 'p': 400, 'q': 150},
        {'id': 'LOAD_NORTH', 'name': 'North Residential', 'bus': 'BUS_NORTH_110', 'p': 250, 'q': 80},
        {'id': 'LOAD_SOUTH', 'name': 'South Industrial', 'bus': 'BUS_SOUTH_110', 'p': 450, 'q': 180},
        {'id': 'LOAD_CENTRAL', 'name': 'Central Load', 'bus': 'BUS_CENTRAL_220', 'p': 200, 'q': 70},
    ])
    # Total charge: 1650 MW
    
    voltage_levels = pd.DataFrame([
        {'id': 'VL_400_CENTRAL', 'name': '400kV Central', 'substation': 'SUB_CENTRAL', 'nominalVoltage': 400000, 'highVoltageLimit': 420000, 'lowVoltageLimit': 380000},
        {'id': 'VL_220_NORTH', 'name': '220kV North', 'substation': 'SUB_NORTH', 'nominalVoltage': 220000, 'highVoltageLimit': 242000, 'lowVoltageLimit': 198000},
        {'id': 'VL_220_SOUTH', 'name': '220kV South', 'substation': 'SUB_SOUTH', 'nominalVoltage': 220000, 'highVoltageLimit': 242000, 'lowVoltageLimit': 198000},
        {'id': 'VL_110_EAST', 'name': '110kV East', 'substation': 'SUB_EAST', 'nominalVoltage': 110000, 'highVoltageLimit': 121000, 'lowVoltageLimit': 99000},
        {'id': 'VL_110_WEST', 'name': '110kV West', 'substation': 'SUB_WEST', 'nominalVoltage': 110000, 'highVoltageLimit': 121000, 'lowVoltageLimit': 99000},
    ])
    
    return {
        'Substations': substations,
        'Buses': buses,
        'Lines': lines,
        'Transformers': transformers,
        'Generators': generators,
        'Loads': loads,
        'Voltage Levels': voltage_levels
    }

def save_to_excel(data, filename):
    """Sauvegarder les données dans un fichier Excel"""
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        for sheet_name, df in data.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]
            
            # Style header
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Auto-adjust columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
    
    total_entities = sum(len(df) for df in data.values())
    print(f"✅ Excel créé: {filename} ({total_entities} entités)")

def create_rdf_small():
    """Créer un fichier RDF petit"""
    rdf_content = """<?xml version="1.0" encoding="UTF-8"?>
<rdf:RDF xmlns:rdf="http://www.w3.org/1999/02/22-rdf-syntax-ns#"
         xmlns:cim="http://iec.ch/TC57/CIM100#"
         xmlns:rdfs="http://www.w3.org/2000/01/rdf-schema#">

    <cim:BaseVoltage rdf:ID="BaseVoltage_220kV">
        <cim:IdentifiedObject.name>220kV</cim:IdentifiedObject.name>
        <cim:BaseVoltage.nominalVoltage>220000</cim:BaseVoltage.nominalVoltage>
    </cim:BaseVoltage>

    <cim:Substation rdf:ID="SUB_A">
        <cim:IdentifiedObject.name>Substation A</cim:IdentifiedObject.name>
    </cim:Substation>

    <cim:Substation rdf:ID="SUB_B">
        <cim:IdentifiedObject.name>Substation B</cim:IdentifiedObject.name>
    </cim:Substation>

    <cim:VoltageLevel rdf:ID="VL_220_A">
        <cim:IdentifiedObject.name>220kV A</cim:IdentifiedObject.name>
        <cim:VoltageLevel.BaseVoltage rdf:resource="#BaseVoltage_220kV"/>
        <cim:VoltageLevel.Substation rdf:resource="#SUB_A"/>
    </cim:VoltageLevel>

    <cim:VoltageLevel rdf:ID="VL_220_B">
        <cim:IdentifiedObject.name>220kV B</cim:IdentifiedObject.name>
        <cim:VoltageLevel.BaseVoltage rdf:resource="#BaseVoltage_220kV"/>
        <cim:VoltageLevel.Substation rdf:resource="#SUB_B"/>
    </cim:VoltageLevel>

    <cim:ConnectivityNode rdf:ID="BUS_A_220">
        <cim:IdentifiedObject.name>Bus A 220kV</cim:IdentifiedObject.name>
        <cim:ConnectivityNode.ConnectivityNodeContainer rdf:resource="#VL_220_A"/>
    </cim:ConnectivityNode>

    <cim:ConnectivityNode rdf:ID="BUS_B_220">
        <cim:IdentifiedObject.name>Bus B 220kV</cim:IdentifiedObject.name>
        <cim:ConnectivityNode.ConnectivityNodeContainer rdf:resource="#VL_220_B"/>
    </cim:ConnectivityNode>

    <cim:GeneratingUnit rdf:ID="GEN_A">
        <cim:IdentifiedObject.name>Generator A</cim:IdentifiedObject.name>
        <cim:GeneratingUnit.maxOperatingP>500</cim:GeneratingUnit.maxOperatingP>
        <cim:GeneratingUnit.minOperatingP>0</cim:GeneratingUnit.minOperatingP>
        <cim:Equipment.EquipmentContainer rdf:resource="#SUB_A"/>
    </cim:GeneratingUnit>

    <cim:EnergyConsumer rdf:ID="LOAD_B">
        <cim:IdentifiedObject.name>Load B</cim:IdentifiedObject.name>
        <cim:EnergyConsumer.p>280</cim:EnergyConsumer.p>
        <cim:EnergyConsumer.q>80</cim:EnergyConsumer.q>
        <cim:Equipment.EquipmentContainer rdf:resource="#SUB_B"/>
    </cim:EnergyConsumer>

    <cim:ACLineSegment rdf:ID="LINE_AB">
        <cim:IdentifiedObject.name>Line A-B</cim:IdentifiedObject.name>
        <cim:Conductor.length>50.0</cim:Conductor.length>
        <cim:ACLineSegment.r>0.05</cim:ACLineSegment.r>
        <cim:ACLineSegment.x>0.15</cim:ACLineSegment.x>
        <cim:ConductingEquipment.BaseVoltage rdf:resource="#BaseVoltage_220kV"/>
    </cim:ACLineSegment>

    <cim:Terminal rdf:ID="T_GEN_A">
        <cim:Terminal.ConductingEquipment rdf:resource="#GEN_A"/>
        <cim:Terminal.ConnectivityNode rdf:resource="#BUS_A_220"/>
    </cim:Terminal>

    <cim:Terminal rdf:ID="T_LOAD_B">
        <cim:Terminal.ConductingEquipment rdf:resource="#LOAD_B"/>
        <cim:Terminal.ConnectivityNode rdf:resource="#BUS_B_220"/>
    </cim:Terminal>

    <cim:Terminal rdf:ID="T_LINE_AB_1">
        <cim:Terminal.ConductingEquipment rdf:resource="#LINE_AB"/>
        <cim:Terminal.ConnectivityNode rdf:resource="#BUS_A_220"/>
    </cim:Terminal>

    <cim:Terminal rdf:ID="T_LINE_AB_2">
        <cim:Terminal.ConductingEquipment rdf:resource="#LINE_AB"/>
        <cim:Terminal.ConnectivityNode rdf:resource="#BUS_B_220"/>
    </cim:Terminal>

</rdf:RDF>"""
    
    return rdf_content

def main():
    """Générer tous les fichiers d'exemple"""
    print("🔌 Génération de fichiers d'exemple pour tests Load Flow")
    print("=" * 60)
    print()
    
    output_dir = '/Users/zakaria/Documents/CIM-SemanticGraph-Platform/examples'
    os.makedirs(output_dir, exist_ok=True)
    
    # Réseau petit
    print("📊 Génération réseau PETIT (2 buses)...")
    small_network = create_small_network()
    save_to_excel(small_network, f'{output_dir}/network-small-2bus.xlsx')
    
    # Réseau moyen
    print("📊 Génération réseau MOYEN (8 buses)...")
    medium_network = create_medium_network()
    save_to_excel(medium_network, f'{output_dir}/network-medium-8bus.xlsx')
    
    # RDF petit
    print("📊 Génération fichier RDF petit...")
    rdf_small = create_rdf_small()
    with open(f'{output_dir}/network-small-2bus.rdf', 'w', encoding='utf-8') as f:
        f.write(rdf_small)
    print(f"✅ RDF créé: {output_dir}/network-small-2bus.rdf")
    
    print()
    print("=" * 60)
    print("✅ Tous les fichiers d'exemple ont été générés!")
    print()
    print("📁 Fichiers créés dans:", output_dir)
    print()
    print("🧪 Testez avec:")
    print("   - network-small-2bus.xlsx (petit)")
    print("   - network-medium-8bus.xlsx (moyen)")
    print("   - network-small-2bus.rdf (RDF)")
    print()

if __name__ == '__main__':
    main()
