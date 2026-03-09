#!/usr/bin/env python3
"""
Generate CIM test files in multiple formats and sizes for the CIM-SemanticGraph-Platform.
Produces:
  - 10 CIM/XML files  (.xml)
  - 10 CIM/RDF files  (.rdf)
  - 10 RDF/XML files  (.rdf  – pure RDF/XML)
  - 10 TURTLE files   (.ttl)
  - 10 Excel files    (.xlsx)
"""

import os
import uuid
import random
import math
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("WARNING: openpyxl not installed – Excel files will be skipped.")
    print("Install with:  pip install openpyxl")

# ── Paths ──────────────────────────────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).parent
CIM_XML_DIR  = SCRIPT_DIR / "cim_xml"
CIM_RDF_DIR  = SCRIPT_DIR / "cim_rdf"
RDF_XML_DIR  = SCRIPT_DIR / "rdf_xml"
TURTLE_DIR   = SCRIPT_DIR / "turtle"
EXCEL_DIR    = SCRIPT_DIR / "excel"

for d in [CIM_XML_DIR, CIM_RDF_DIR, RDF_XML_DIR, TURTLE_DIR, EXCEL_DIR]:
    d.mkdir(exist_ok=True)

# ── Namespaces ─────────────────────────────────────────────────────────────────
CIM_NS   = "http://iec.ch/TC57/CIM100#"
RDF_NS   = "http://www.w3.org/1999/02/22-rdf-syntax-ns#"
RDFS_NS  = "http://www.w3.org/2000/01/rdf-schema#"
OWL_NS   = "http://www.w3.org/2002/07/owl#"
MD_NS    = "http://iec.ch/TC57/61970-552/ModelDescription/1#"
EQ_NS    = "http://iec.ch/TC57/2013/CIM-schema-cim16/EquipmentCore#"

# ── Size configs ───────────────────────────────────────────────────────────────
# (label, n_buses, n_lines, n_gens, n_loads, n_trafos)
SIZES = [
    ("tiny_1bus",      1,   0, 1,  1,  0),
    ("micro_2bus",     2,   1, 1,  1,  1),
    ("small_4bus",     4,   4, 2,  2,  1),
    ("small_6bus",     6,   6, 2,  3,  2),
    ("small_8bus",     8,   8, 3,  4,  2),
    ("medium_12bus",  12,  14, 4,  6,  3),
    ("medium_20bus",  20,  24, 6,  8,  4),
    ("medium_30bus",  30,  36, 8, 12,  5),
    ("large_50bus",   50,  65,12, 20,  8),
    ("large_100bus", 100, 130,20, 40, 12),
]

# ── Helpers ────────────────────────────────────────────────────────────────────
def uid(): return "_" + str(uuid.uuid4()).replace("-","")

def base_voltage(n):
    """Pick a realistic base voltage kV depending on network size."""
    if n <= 4:   return 10.5
    if n <= 10:  return 20.0
    if n <= 30:  return 110.0
    return 220.0

def gen_topology(n_buses, n_lines):
    """Return list of (from_bus_idx, to_bus_idx) for a connected network."""
    edges = []
    # Spanning tree first
    for i in range(1, n_buses):
        j = random.randint(0, i-1)
        edges.append((j, i))
    # Extra lines
    extras = max(0, n_lines - (n_buses - 1))
    for _ in range(extras):
        a = random.randint(0, n_buses - 1)
        b = random.randint(0, n_buses - 1)
        if a != b:
            edges.append((a, b))
    return edges

# ═══════════════════════════════════════════════════════════════════════════════
# CIM/XML  (IEC 61970 CIM/XML header, rdf:RDF body)
# ═══════════════════════════════════════════════════════════════════════════════
CIM_XML_HEADER = """\
<?xml version="1.0" encoding="UTF-8"?>
<!-- CIM/XML  IEC 61970-552  –  {label} -->
<rdf:RDF
  xmlns:rdf="{RDF}"
  xmlns:rdfs="{RDFS}"
  xmlns:cim="{CIM}"
  xmlns:md="{MD}">

  <md:FullModel rdf:about="urn:uuid:{model_id}">
    <md:Model.description>CIM test network – {label}</md:Model.description>
    <md:Model.version>1</md:Model.version>
    <md:Model.profile>http://iec.ch/TC57/61970-552/ModelDescription/1</md:Model.profile>
  </md:FullModel>
"""

def build_cim_xml(label, buses, edges, gens, loads, trafos):
    bv_kv = base_voltage(len(buses))
    bv_id = uid()
    lines = [CIM_XML_HEADER.format(
        label=label, model_id=uid().lstrip("_"),
        RDF=RDF_NS, RDFS=RDFS_NS, CIM=CIM_NS, MD=MD_NS)]

    # BaseVoltage
    lines.append(f'  <cim:BaseVoltage rdf:ID="{bv_id}">')
    lines.append(f'    <cim:IdentifiedObject.name>BV_{bv_kv}kV</cim:IdentifiedObject.name>')
    lines.append(f'    <cim:BaseVoltage.nominalVoltage>{bv_kv}</cim:BaseVoltage.nominalVoltage>')
    lines.append( '  </cim:BaseVoltage>')

    # Substation
    sub_id = uid()
    lines.append(f'  <cim:Substation rdf:ID="{sub_id}">')
    lines.append(f'    <cim:IdentifiedObject.name>SS_{label}</cim:IdentifiedObject.name>')
    lines.append( '  </cim:Substation>')

    # VoltageLevel
    vl_id = uid()
    lines.append(f'  <cim:VoltageLevel rdf:ID="{vl_id}">')
    lines.append(f'    <cim:IdentifiedObject.name>VL_{label}</cim:IdentifiedObject.name>')
    lines.append(f'    <cim:VoltageLevel.Substation rdf:resource="#{sub_id}"/>')
    lines.append(f'    <cim:VoltageLevel.BaseVoltage rdf:resource="#{bv_id}"/>')
    lines.append( '  </cim:VoltageLevel>')

    # ConnectivityNodes / BusbarSections
    cn_ids = []
    for i, b in enumerate(buses):
        cn = uid()
        cn_ids.append(cn)
        lines.append(f'  <cim:ConnectivityNode rdf:ID="{cn}">')
        lines.append(f'    <cim:IdentifiedObject.name>{b}</cim:IdentifiedObject.name>')
        lines.append(f'    <cim:ConnectivityNode.ConnectivityNodeContainer rdf:resource="#{vl_id}"/>')
        lines.append( '  </cim:ConnectivityNode>')
        bb = uid()
        lines.append(f'  <cim:BusbarSection rdf:ID="{bb}">')
        lines.append(f'    <cim:IdentifiedObject.name>BB_{b}</cim:IdentifiedObject.name>')
        t1 = uid()
        lines.append(f'    <cim:ConductingEquipment.BaseVoltage rdf:resource="#{bv_id}"/>')
        lines.append( '  </cim:BusbarSection>')
        # Terminal for busbar
        lines.append(f'  <cim:Terminal rdf:ID="{t1}">')
        lines.append(f'    <cim:Terminal.ConductingEquipment rdf:resource="#{bb}"/>')
        lines.append(f'    <cim:Terminal.ConnectivityNode rdf:resource="#{cn}"/>')
        lines.append( '  </cim:Terminal>')

    # ACLineSegments
    for idx, (a, b) in enumerate(edges):
        ln = uid()
        r  = round(random.uniform(0.01, 2.5), 4)
        x  = round(random.uniform(0.05, 5.0), 4)
        bch = round(random.uniform(1e-6, 1e-4), 8)
        lines.append(f'  <cim:ACLineSegment rdf:ID="{ln}">')
        lines.append(f'    <cim:IdentifiedObject.name>L{idx+1}_{buses[a]}_{buses[b]}</cim:IdentifiedObject.name>')
        lines.append(f'    <cim:Conductor.length>{round(random.uniform(1,200),1)}</cim:Conductor.length>')
        lines.append(f'    <cim:ACLineSegment.r>{r}</cim:ACLineSegment.r>')
        lines.append(f'    <cim:ACLineSegment.x>{x}</cim:ACLineSegment.x>')
        lines.append(f'    <cim:ACLineSegment.bch>{bch}</cim:ACLineSegment.bch>')
        lines.append(f'    <cim:ConductingEquipment.BaseVoltage rdf:resource="#{bv_id}"/>')
        lines.append( '  </cim:ACLineSegment>')
        for end, cn_idx in enumerate([a, b]):
            t = uid()
            lines.append(f'  <cim:Terminal rdf:ID="{t}">')
            lines.append(f'    <cim:Terminal.ConductingEquipment rdf:resource="#{ln}"/>')
            lines.append(f'    <cim:Terminal.ConnectivityNode rdf:resource="#{cn_ids[cn_idx]}"/>')
            lines.append(f'    <cim:ACDCTerminal.sequenceNumber>{end+1}</cim:ACDCTerminal.sequenceNumber>')
            lines.append( '  </cim:Terminal>')

    # Generators
    for i, g in enumerate(gens):
        bus_idx = i % len(buses)
        gn = uid()
        max_p = round(random.uniform(50,500), 1)
        lines.append(f'  <cim:GeneratingUnit rdf:ID="{gn}">')
        lines.append(f'    <cim:IdentifiedObject.name>GEN_{i+1}_{buses[bus_idx]}</cim:IdentifiedObject.name>')
        lines.append(f'    <cim:GeneratingUnit.maxOperatingP>{max_p}</cim:GeneratingUnit.maxOperatingP>')
        lines.append(f'    <cim:GeneratingUnit.minOperatingP>{round(max_p*0.1,1)}</cim:GeneratingUnit.minOperatingP>')
        lines.append(f'    <cim:GeneratingUnit.nominalP>{round(max_p*0.8,1)}</cim:GeneratingUnit.nominalP>')
        lines.append( '  </cim:GeneratingUnit>')
        sm = uid()
        lines.append(f'  <cim:SynchronousMachine rdf:ID="{sm}">')
        lines.append(f'    <cim:IdentifiedObject.name>SM_{i+1}</cim:IdentifiedObject.name>')
        lines.append(f'    <cim:SynchronousMachine.GeneratingUnit rdf:resource="#{gn}"/>')
        lines.append(f'    <cim:ConductingEquipment.BaseVoltage rdf:resource="#{bv_id}"/>')
        lines.append(f'    <cim:RegulatingCondEq.RegulatingControl rdf:resource="#{gn}"/>')
        lines.append( '  </cim:SynchronousMachine>')
        t = uid()
        lines.append(f'  <cim:Terminal rdf:ID="{t}">')
        lines.append(f'    <cim:Terminal.ConductingEquipment rdf:resource="#{sm}"/>')
        lines.append(f'    <cim:Terminal.ConnectivityNode rdf:resource="#{cn_ids[bus_idx]}"/>')
        lines.append( '  </cim:Terminal>')

    # Loads
    for i, l in enumerate(loads):
        bus_idx = (i + len(gens)) % len(buses)
        ec = uid()
        p = round(random.uniform(10, 300), 1)
        lines.append(f'  <cim:EnergyConsumer rdf:ID="{ec}">')
        lines.append(f'    <cim:IdentifiedObject.name>LOAD_{i+1}_{buses[bus_idx]}</cim:IdentifiedObject.name>')
        lines.append(f'    <cim:EnergyConsumer.p>{p}</cim:EnergyConsumer.p>')
        lines.append(f'    <cim:EnergyConsumer.q>{round(p*random.uniform(0.1,0.4),1)}</cim:EnergyConsumer.q>')
        lines.append(f'    <cim:ConductingEquipment.BaseVoltage rdf:resource="#{bv_id}"/>')
        lines.append( '  </cim:EnergyConsumer>')
        t = uid()
        lines.append(f'  <cim:Terminal rdf:ID="{t}">')
        lines.append(f'    <cim:Terminal.ConductingEquipment rdf:resource="#{ec}"/>')
        lines.append(f'    <cim:Terminal.ConnectivityNode rdf:resource="#{cn_ids[bus_idx]}"/>')
        lines.append( '  </cim:Terminal>')

    # Transformers
    for i in range(trafos):
        pt = uid()
        lines.append(f'  <cim:PowerTransformer rdf:ID="{pt}">')
        lines.append(f'    <cim:IdentifiedObject.name>TRAFO_{i+1}</cim:IdentifiedObject.name>')
        lines.append( '  </cim:PowerTransformer>')
        for winding in range(2):
            ptw = uid()
            lines.append(f'  <cim:PowerTransformerEnd rdf:ID="{ptw}">')
            lines.append(f'    <cim:IdentifiedObject.name>TRAFO_{i+1}_W{winding+1}</cim:IdentifiedObject.name>')
            lines.append(f'    <cim:PowerTransformerEnd.PowerTransformer rdf:resource="#{pt}"/>')
            lines.append(f'    <cim:TransformerEnd.endNumber>{winding+1}</cim:TransformerEnd.endNumber>')
            lines.append(f'    <cim:PowerTransformerEnd.ratedS>{round(random.uniform(100,500),1)}</cim:PowerTransformerEnd.ratedS>')
            lines.append(f'    <cim:PowerTransformerEnd.ratedU>{bv_kv if winding==0 else bv_kv/math.sqrt(3)}</cim:PowerTransformerEnd.ratedU>')
            lines.append(f'    <cim:PowerTransformerEnd.r>{round(random.uniform(0.1,2.0),4)}</cim:PowerTransformerEnd.r>')
            lines.append(f'    <cim:PowerTransformerEnd.x>{round(random.uniform(1.0,15.0),4)}</cim:PowerTransformerEnd.x>')
            lines.append( '  </cim:PowerTransformerEnd>')
            bus_idx = (i*2+winding) % len(buses)
            t = uid()
            lines.append(f'  <cim:Terminal rdf:ID="{t}">')
            lines.append(f'    <cim:Terminal.ConductingEquipment rdf:resource="#{pt}"/>')
            lines.append(f'    <cim:Terminal.ConnectivityNode rdf:resource="#{cn_ids[bus_idx]}"/>')
            lines.append(f'    <cim:ACDCTerminal.sequenceNumber>{winding+1}</cim:ACDCTerminal.sequenceNumber>')
            lines.append( '  </cim:Terminal>')

    lines.append("</rdf:RDF>")
    return "\n".join(lines)

# ═══════════════════════════════════════════════════════════════════════════════
# TURTLE
# ═══════════════════════════════════════════════════════════════════════════════
TURTLE_PREFIXES = """\
@prefix rdf:  <http://www.w3.org/1999/02/22-rdf-syntax-ns#> .
@prefix rdfs: <http://www.w3.org/2000/01/rdf-schema#> .
@prefix owl:  <http://www.w3.org/2002/07/owl#> .
@prefix cim:  <http://iec.ch/TC57/CIM100#> .
@prefix md:   <http://iec.ch/TC57/61970-552/ModelDescription/1#> .
@prefix xsd:  <http://www.w3.org/2001/XMLSchema#> .

"""

def build_turtle(label, buses, edges, gens, loads, trafos):
    bv_kv  = base_voltage(len(buses))
    bv_uri = f"cim:BV_{label}"
    sub_uri = f"cim:SS_{label}"
    vl_uri  = f"cim:VL_{label}"

    lines = [TURTLE_PREFIXES]
    lines.append(f'# CIM Turtle – {label}  ({len(buses)} buses, {len(edges)} lines)\n')

    lines.append(f'{bv_uri} a cim:BaseVoltage ;')
    lines.append(f'    cim:IdentifiedObject.name "BV_{bv_kv}kV" ;')
    lines.append(f'    cim:BaseVoltage.nominalVoltage {bv_kv} .')

    lines.append(f'\n{sub_uri} a cim:Substation ;')
    lines.append(f'    cim:IdentifiedObject.name "SS_{label}" .')

    lines.append(f'\n{vl_uri} a cim:VoltageLevel ;')
    lines.append(f'    cim:IdentifiedObject.name "VL_{label}" ;')
    lines.append(f'    cim:VoltageLevel.Substation {sub_uri} ;')
    lines.append(f'    cim:VoltageLevel.BaseVoltage {bv_uri} .')

    cn_uris = []
    for i, b in enumerate(buses):
        cn_uri = f"cim:CN_{label}_{i}"
        cn_uris.append(cn_uri)
        bb_uri = f"cim:BB_{label}_{i}"
        lines.append(f'\n{cn_uri} a cim:ConnectivityNode ;')
        lines.append(f'    cim:IdentifiedObject.name "{b}" ;')
        lines.append(f'    cim:ConnectivityNode.ConnectivityNodeContainer {vl_uri} .')
        lines.append(f'\n{bb_uri} a cim:BusbarSection ;')
        lines.append(f'    cim:IdentifiedObject.name "BB_{b}" ;')
        lines.append(f'    cim:ConductingEquipment.BaseVoltage {bv_uri} .')
        t_uri = f"cim:T_BB_{label}_{i}"
        lines.append(f'\n{t_uri} a cim:Terminal ;')
        lines.append(f'    cim:Terminal.ConductingEquipment {bb_uri} ;')
        lines.append(f'    cim:Terminal.ConnectivityNode {cn_uri} .')

    for idx, (a, b) in enumerate(edges):
        ln_uri = f"cim:L_{label}_{idx}"
        r  = round(random.uniform(0.01, 2.5), 4)
        x  = round(random.uniform(0.05, 5.0), 4)
        bch = round(random.uniform(1e-6, 1e-4), 8)
        lines.append(f'\n{ln_uri} a cim:ACLineSegment ;')
        lines.append(f'    cim:IdentifiedObject.name "L{idx+1}_{buses[a]}_{buses[b]}" ;')
        lines.append(f'    cim:ACLineSegment.r {r} ;')
        lines.append(f'    cim:ACLineSegment.x {x} ;')
        lines.append(f'    cim:ACLineSegment.bch {bch} ;')
        lines.append(f'    cim:ConductingEquipment.BaseVoltage {bv_uri} .')
        for end, cn_idx in enumerate([a, b]):
            t_uri = f"cim:T_L_{label}_{idx}_{end}"
            lines.append(f'\n{t_uri} a cim:Terminal ;')
            lines.append(f'    cim:Terminal.ConductingEquipment {ln_uri} ;')
            lines.append(f'    cim:Terminal.ConnectivityNode {cn_uris[cn_idx]} ;')
            lines.append(f'    cim:ACDCTerminal.sequenceNumber {end+1} .')

    for i in range(gens):
        bus_idx = i % len(buses)
        gn_uri = f"cim:GEN_{label}_{i}"
        sm_uri = f"cim:SM_{label}_{i}"
        max_p  = round(random.uniform(50, 500), 1)
        lines.append(f'\n{gn_uri} a cim:GeneratingUnit ;')
        lines.append(f'    cim:IdentifiedObject.name "GEN_{i+1}" ;')
        lines.append(f'    cim:GeneratingUnit.maxOperatingP {max_p} ;')
        lines.append(f'    cim:GeneratingUnit.minOperatingP {round(max_p*0.1,1)} ;')
        lines.append(f'    cim:GeneratingUnit.nominalP {round(max_p*0.8,1)} .')
        lines.append(f'\n{sm_uri} a cim:SynchronousMachine ;')
        lines.append(f'    cim:IdentifiedObject.name "SM_{i+1}" ;')
        lines.append(f'    cim:SynchronousMachine.GeneratingUnit {gn_uri} ;')
        lines.append(f'    cim:ConductingEquipment.BaseVoltage {bv_uri} .')
        t_uri = f"cim:T_GEN_{label}_{i}"
        lines.append(f'\n{t_uri} a cim:Terminal ;')
        lines.append(f'    cim:Terminal.ConductingEquipment {sm_uri} ;')
        lines.append(f'    cim:Terminal.ConnectivityNode {cn_uris[bus_idx]} .')

    for i in range(loads):
        bus_idx = (i + gens) % len(buses)
        ec_uri  = f"cim:LOAD_{label}_{i}"
        p = round(random.uniform(10, 300), 1)
        lines.append(f'\n{ec_uri} a cim:EnergyConsumer ;')
        lines.append(f'    cim:IdentifiedObject.name "LOAD_{i+1}" ;')
        lines.append(f'    cim:EnergyConsumer.p {p} ;')
        lines.append(f'    cim:EnergyConsumer.q {round(p*random.uniform(0.1,0.4),1)} ;')
        lines.append(f'    cim:ConductingEquipment.BaseVoltage {bv_uri} .')
        t_uri = f"cim:T_LOAD_{label}_{i}"
        lines.append(f'\n{t_uri} a cim:Terminal ;')
        lines.append(f'    cim:Terminal.ConductingEquipment {ec_uri} ;')
        lines.append(f'    cim:Terminal.ConnectivityNode {cn_uris[bus_idx]} .')

    for i in range(trafos):
        pt_uri = f"cim:PT_{label}_{i}"
        lines.append(f'\n{pt_uri} a cim:PowerTransformer ;')
        lines.append(f'    cim:IdentifiedObject.name "TRAFO_{i+1}" .')
        for winding in range(2):
            ptw_uri = f"cim:PTW_{label}_{i}_{winding}"
            bus_idx = (i*2+winding) % len(buses)
            lines.append(f'\n{ptw_uri} a cim:PowerTransformerEnd ;')
            lines.append(f'    cim:PowerTransformerEnd.PowerTransformer {pt_uri} ;')
            lines.append(f'    cim:TransformerEnd.endNumber {winding+1} ;')
            lines.append(f'    cim:PowerTransformerEnd.ratedS {round(random.uniform(100,500),1)} ;')
            lines.append(f'    cim:PowerTransformerEnd.ratedU {bv_kv if winding==0 else round(bv_kv/math.sqrt(3),2)} .')
            t_uri = f"cim:T_PT_{label}_{i}_{winding}"
            lines.append(f'\n{t_uri} a cim:Terminal ;')
            lines.append(f'    cim:Terminal.ConductingEquipment {pt_uri} ;')
            lines.append(f'    cim:Terminal.ConnectivityNode {cn_uris[bus_idx]} ;')
            lines.append(f'    cim:ACDCTerminal.sequenceNumber {winding+1} .')

    return "\n".join(lines) + "\n"

# ═══════════════════════════════════════════════════════════════════════════════
# Excel
# ═══════════════════════════════════════════════════════════════════════════════
HDR_FILL  = PatternFill("solid", fgColor="1E3A5F")
HDR_FONT  = Font(bold=True, color="FFFFFF", size=10)
HDR_ALIGN = Alignment(horizontal="center", vertical="center")
ALT_FILL  = PatternFill("solid", fgColor="EEF2FF")

def style_header(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill  = HDR_FILL
        cell.font  = HDR_FONT
        cell.alignment = HDR_ALIGN

def style_alt(ws, row, cols):
    if row % 2 == 0:
        for c in range(1, cols+1):
            ws.cell(row=row, column=c).fill = ALT_FILL

def write_sheet_buses(ws, buses, bv_kv):
    headers = ["Bus ID","Bus Name","Voltage (kV)","Vmin (pu)","Vmax (pu)","Type"]
    ws.append(headers); style_header(ws, 1, len(headers))
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 22
    for i, b in enumerate(buses):
        row = [f"BUS_{i+1:03d}", b, bv_kv, 0.95, 1.05,
               "SLACK" if i == 0 else random.choice(["PQ","PV"])]
        ws.append(row); style_alt(ws, i+2, len(headers))

def write_sheet_lines(ws, buses, edges):
    headers = ["Line ID","From Bus","To Bus","R (pu)","X (pu)","B (pu)","Length (km)","Rating (MVA)"]
    ws.append(headers); style_header(ws, 1, len(headers))
    ws.column_dimensions["B"].width = 22; ws.column_dimensions["C"].width = 22
    for idx, (a, b) in enumerate(edges):
        row = [f"L{idx+1:03d}", buses[a], buses[b],
               round(random.uniform(0.001,0.05),5),
               round(random.uniform(0.005,0.1),5),
               round(random.uniform(1e-4,5e-3),6),
               round(random.uniform(1,200),1),
               round(random.uniform(50,400),0)]
        ws.append(row); style_alt(ws, idx+2, len(headers))

def write_sheet_generators(ws, buses, n_gens):
    headers = ["Gen ID","Bus","Fuel Type","Pmax (MW)","Pmin (MW)","Pset (MW)","Qmax (MVAr)","Qmin (MVAr)","V setpoint (pu)"]
    ws.append(headers); style_header(ws, 1, len(headers))
    fuels = ["GAS","COAL","NUCLEAR","HYDRO","WIND","SOLAR"]
    for i in range(n_gens):
        bus_idx = i % len(buses)
        max_p = round(random.uniform(50,500),1)
        row = [f"G{i+1:03d}", buses[bus_idx], random.choice(fuels),
               max_p, round(max_p*0.1,1), round(max_p*0.8,1),
               round(max_p*0.5,1), round(-max_p*0.3,1), 1.0]
        ws.append(row); style_alt(ws, i+2, len(headers))

def write_sheet_loads(ws, buses, n_loads):
    headers = ["Load ID","Bus","P (MW)","Q (MVAr)","Power Factor","Type"]
    ws.append(headers); style_header(ws, 1, len(headers))
    types = ["Residential","Industrial","Commercial","Agricultural"]
    for i in range(n_loads):
        bus_idx = (i + 1) % len(buses)
        p = round(random.uniform(10,300),1)
        q = round(p*random.uniform(0.1,0.4),1)
        row = [f"LD{i+1:03d}", buses[bus_idx], p, q,
               round(p/math.sqrt(p**2+q**2),3), random.choice(types)]
        ws.append(row); style_alt(ws, i+2, len(headers))

def write_sheet_transformers(ws, buses, n_trafos, bv_kv):
    headers = ["Trafo ID","Primary Bus","Secondary Bus","Rated S (MVA)","V1 (kV)","V2 (kV)","R (pu)","X (pu)","Tap ratio"]
    ws.append(headers); style_header(ws, 1, len(headers))
    for i in range(n_trafos):
        p_bus = buses[i % len(buses)]
        s_bus = buses[(i+1) % len(buses)]
        row = [f"TR{i+1:03d}", p_bus, s_bus,
               round(random.uniform(100,500),1), bv_kv, round(bv_kv/10,1),
               round(random.uniform(0.001,0.01),5),
               round(random.uniform(0.05,0.15),5),
               round(random.uniform(0.95,1.05),4)]
        ws.append(row); style_alt(ws, i+2, len(headers))

def build_excel(label, buses, edges, gens, loads, trafos, out_path):
    wb = openpyxl.Workbook()
    bv_kv = base_voltage(len(buses))

    # Cover sheet
    ws0 = wb.active; ws0.title = "Info"
    ws0["A1"] = "CIM Network Test File"; ws0["A1"].font = Font(bold=True, size=14, color="1E3A5F")
    ws0["A2"] = f"Network: {label}"
    ws0["A3"] = f"Buses: {len(buses)}  |  Lines: {len(edges)}  |  Generators: {gens}  |  Loads: {loads}  |  Transformers: {trafos}"
    ws0["A4"] = f"Base Voltage: {bv_kv} kV"
    ws0["A5"] = "Generated by CIM-SemanticGraph-Platform test generator"
    ws0.column_dimensions["A"].width = 60

    ws1 = wb.create_sheet("Buses");         write_sheet_buses(ws1, buses, bv_kv)
    ws2 = wb.create_sheet("Lines");         write_sheet_lines(ws2, buses, edges)
    ws3 = wb.create_sheet("Generators");    write_sheet_generators(ws3, buses, gens)
    ws4 = wb.create_sheet("Loads");         write_sheet_loads(ws4, buses, loads)
    ws5 = wb.create_sheet("Transformers");  write_sheet_transformers(ws5, buses, trafos, bv_kv)

    wb.save(out_path)

# ═══════════════════════════════════════════════════════════════════════════════
# Main generation loop
# ═══════════════════════════════════════════════════════════════════════════════
def bus_names(n):
    return [f"BUS_{i+1:03d}" for i in range(n)]

random.seed(42)
generated = {"cim_xml":[], "cim_rdf":[], "rdf_xml":[], "turtle":[], "excel":[]}

for label, n_buses, n_lines, n_gens, n_loads, n_trafos in SIZES:
    buses  = bus_names(n_buses)
    edges  = gen_topology(n_buses, n_lines) if n_buses > 1 else []
    gens   = list(range(n_gens))
    loads  = list(range(n_loads))

    xml_body = build_cim_xml(label, buses, edges, gens, loads, n_trafos)
    ttl_body = build_turtle(label, buses, edges, n_gens, n_loads, n_trafos)

    # 1. CIM/XML  (.xml)
    p = CIM_XML_DIR / f"{label}.xml"
    p.write_text(xml_body, encoding="utf-8")
    generated["cim_xml"].append((p.name, p.stat().st_size))

    # 2. CIM/RDF  (.rdf)  – same content, different extension (common in CIM world)
    p = CIM_RDF_DIR / f"{label}.rdf"
    p.write_text(xml_body, encoding="utf-8")
    generated["cim_rdf"].append((p.name, p.stat().st_size))

    # 3. RDF/XML  (.rdf) – identical serialisation, renamed to emphasise W3C format
    p = RDF_XML_DIR / f"{label}_rdfxml.rdf"
    p.write_text(xml_body, encoding="utf-8")
    generated["rdf_xml"].append((p.name, p.stat().st_size))

    # 4. TURTLE  (.ttl)
    p = TURTLE_DIR / f"{label}.ttl"
    p.write_text(ttl_body, encoding="utf-8")
    generated["turtle"].append((p.name, p.stat().st_size))

    # 5. Excel  (.xlsx)
    if HAS_OPENPYXL:
        p = EXCEL_DIR / f"{label}.xlsx"
        build_excel(label, buses, edges, n_gens, n_loads, n_trafos, p)
        generated["excel"].append((p.name, p.stat().st_size))

# ── Summary ────────────────────────────────────────────────────────────────────
print("\n" + "═"*60)
print("  CIM Test Files Generated")
print("═"*60)
for fmt, files in generated.items():
    if not files: continue
    print(f"\n  [{fmt.upper()}]  →  {list(SIZES[0])[0]}  …  format")
    for name, size in files:
        kb = size / 1024
        bar = "█" * min(int(kb/2), 40)
        print(f"    {name:<38}  {kb:7.1f} KB  {bar}")
print(f"\n  Total files: {sum(len(v) for v in generated.values())}")
print("═"*60 + "\n")
