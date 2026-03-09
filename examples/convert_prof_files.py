#!/usr/bin/env python3
"""
Converter: Professor's PyPSA/pandapower files → CIM Platform Excel format
============================================================================
Usage:
    python3 convert_prof_files.py

Converts BOTH subnetworks (3 and 4) from the "cim test" folder into:
  - examples/converted/subnetwork_3_merged.xlsx   (uploadable directly)
  - examples/converted/subnetwork_4_merged.xlsx
  - examples/converted/subnetwork_3.rdf           (CIM/RDF alternative)
  - examples/converted/subnetwork_4.rdf
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path
import uuid, math

# ── Paths ──────────────────────────────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).parent
CIM_TEST   = SCRIPT_DIR / "cim test"
OUT_DIR    = SCRIPT_DIR / "converted"
OUT_DIR.mkdir(exist_ok=True)

SUBNETWORKS = {
    "3": CIM_TEST / "subnetwork_valid_3_scrambled",
    "4": CIM_TEST / "subnetwork_valid_4_scrambled",
}

# ── Styles ─────────────────────────────────────────────────────────────────────
HDR_FILL  = PatternFill("solid", fgColor="1E3A5F")
HDR_FONT  = Font(bold=True, color="FFFFFF", size=10)
ALT_FILL  = PatternFill("solid", fgColor="EEF2FF")
HDR_ALIGN = Alignment(horizontal="center")

def style_header(ws, n_cols):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill  = HDR_FILL
        cell.font  = HDR_FONT
        cell.alignment = HDR_ALIGN

def style_row(ws, row_num, n_cols):
    if row_num % 2 == 0:
        for c in range(1, n_cols + 1):
            ws.cell(row=row_num, column=c).fill = ALT_FILL

# ── Read helper ────────────────────────────────────────────────────────────────
def read_sheet(path):
    """Return (headers_list, rows_list_of_dicts)."""
    wb  = openpyxl.load_workbook(path, read_only=True)
    ws  = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not all_rows:
        return [], []
    headers = [str(h).strip() if h is not None else "" for h in all_rows[0]]
    rows    = []
    for r in all_rows[1:]:
        d = {headers[i]: r[i] for i in range(len(headers)) if i < len(r)}
        if any(v is not None for v in d.values()):
            rows.append(d)
    return headers, rows

# ── Conversion logic ───────────────────────────────────────────────────────────
def convert_subnetwork(num: str, folder: Path):
    print(f"\n{'─'*60}")
    print(f"  Converting subnetwork {num}  ({folder.name})")
    print(f"{'─'*60}")

    # Find files
    nodes_file  = next(folder.glob("*nodes*.xlsx"),  None)
    lines_file  = next(folder.glob("*lines*.xlsx"),  None)
    consumers_f = next(folder.glob("*consumers*.xlsx"), None)
    trafos_file = next(folder.glob("*transformers*.xlsx"), None)

    for label, f in [("nodes", nodes_file), ("lines", lines_file),
                     ("consumers", consumers_f), ("transformers", trafos_file)]:
        if f is None:
            print(f"  WARNING: no {label} file found in {folder}")

    # Read data
    _, nodes     = read_sheet(nodes_file)   if nodes_file  else ([], [])
    _, lines     = read_sheet(lines_file)   if lines_file  else ([], [])
    _, consumers = read_sheet(consumers_f)  if consumers_f else ([], [])
    _, trafos    = read_sheet(trafos_file)  if trafos_file else ([], [])

    print(f"  Nodes:     {len(nodes)}")
    print(f"  Lines:     {len(lines)}")
    print(f"  Consumers: {len(consumers)}")
    print(f"  Transformers: {len(trafos)}")

    # ── Build merged Excel ─────────────────────────────────────────────────────
    wb = openpyxl.Workbook()

    # INFO sheet
    ws0 = wb.active; ws0.title = "Info"
    ws0["A1"] = f"CIM Network – Subnetwork {num} (from Prof)"
    ws0["A1"].font = Font(bold=True, size=13, color="1E3A5F")
    ws0["A2"] = f"Buses: {len(nodes)}  |  Lines: {len(lines)}  |  Consumers: {len(consumers)}  |  Transformers: {len(trafos)}"
    ws0["A3"] = "Converted from PyPSA/pandapower format for CIM-SemanticGraph-Platform"
    ws0.column_dimensions["A"].width = 65

    # ── BUSES sheet ───────────────────────────────────────────────────────────
    ws_buses = wb.create_sheet("Buses")
    bus_headers = ["id", "name", "voltage (kV)", "type", "x", "y"]
    ws_buses.append(bus_headers)
    style_header(ws_buses, len(bus_headers))
    for i, n in enumerate(nodes):
        row = [
            n.get("id_") or n.get("id", f"BUS_{i+1}"),
            n.get("id_") or n.get("id", f"BUS_{i+1}"),
            round(float(n["v_nom"]), 4) if n.get("v_nom") else 0.4,  # kV (PyPSA v_nom already in kV)
            n.get("type", ""),
            round(float(n["x_plot"]), 3) if n.get("x_plot") else "",
            round(float(n["y_plot"]), 3) if n.get("y_plot") else "",
        ]
        ws_buses.append(row)
        style_row(ws_buses, i + 2, len(bus_headers))
    ws_buses.column_dimensions["A"].width = 16
    ws_buses.column_dimensions["B"].width = 16
    ws_buses.column_dimensions["C"].width = 14

    # ── LINES sheet ───────────────────────────────────────────────────────────
    ws_lines = wb.create_sheet("Lines")
    line_headers = ["id", "from", "to", "v_nom (kV)", "length (km)", "r (Ohm/km)", "s_nom (MVA)"]
    ws_lines.append(line_headers)
    style_header(ws_lines, len(line_headers))
    for i, l in enumerate(lines):
        row = [
            l.get("id_") or l.get("id", f"L{i+1}"),
            l.get("bus0", ""),
            l.get("bus1", ""),
            round(float(l["v_nom"]), 4) if l.get("v_nom") else "",  # kV (PyPSA v_nom already in kV)
            round(float(l["length"]), 4) if l.get("length") else "",
            round(float(l["r"]), 6) if l.get("r") else "",
            round(float(l["s_nom"]), 3) if l.get("s_nom") else "",
        ]
        ws_lines.append(row)
        style_row(ws_lines, i + 2, len(line_headers))
    ws_lines.column_dimensions["B"].width = 16
    ws_lines.column_dimensions["C"].width = 16

    # ── LOADS (consumers) sheet ───────────────────────────────────────────────
    ws_loads = wb.create_sheet("Loads")
    load_headers = ["id", "bus", "p (MW)", "q (MVAr)", "profile", "pv_kw", "sub_grid"]
    ws_loads.append(load_headers)
    style_header(ws_loads, len(load_headers))
    for i, c in enumerate(consumers):
        # demand_power is annual consumption in kWh → convert to average MW: kWh/year ÷ 8760 h/year ÷ 1000 kW/MW
        dp = c.get("demand_power")
        p_mw = round(float(dp) / 8_760_000.0, 6) if dp else ""
        q_mvar = round(float(p_mw) * 0.2, 6) if p_mw != "" else ""  # assume 0.2 power factor offset
        # pv capacity
        pv_raw = c.get("pv", "")
        pv_kw = ""
        if pv_raw and pv_raw != "[]" and isinstance(pv_raw, str) and "pdc0" in pv_raw:
            try:
                import ast
                pv_list = ast.literal_eval(pv_raw)
                pv_kw = sum(x.get("pdc0", 0) for x in pv_list)
            except Exception:
                pass
        row = [
            c.get("id_") or c.get("id", f"LOAD_{i+1}"),
            c.get("bus0", ""),
            p_mw,
            q_mvar,
            c.get("profile", ""),
            pv_kw,
            c.get("sub_grid", ""),
        ]
        ws_loads.append(row)
        style_row(ws_loads, i + 2, len(load_headers))
    ws_loads.column_dimensions["A"].width = 40
    ws_loads.column_dimensions["B"].width = 16

    # ── TRANSFORMERS sheet ────────────────────────────────────────────────────
    ws_trafos = wb.create_sheet("Transformers")
    trafo_headers = ["id", "hv_bus", "lv_bus", "v0 (kV)", "v1 (kV)", "r", "g", "b", "s_nom (MVA)"]
    ws_trafos.append(trafo_headers)
    style_header(ws_trafos, len(trafo_headers))
    for i, t in enumerate(trafos):
        row = [
            t.get("id_") or t.get("id", f"TR{i+1}"),
            t.get("bus0", ""),
            t.get("bus1", ""),
            t.get("v0", ""),
            t.get("v1", ""),
            t.get("r", ""),
            t.get("g", ""),
            t.get("b", ""),
            t.get("s_nom", ""),
        ]
        ws_trafos.append(row)
        style_row(ws_trafos, i + 2, len(trafo_headers))
    ws_trafos.column_dimensions["B"].width = 16
    ws_trafos.column_dimensions["C"].width = 16

    out_xlsx = OUT_DIR / f"subnetwork_{num}_merged.xlsx"
    wb.save(out_xlsx)
    print(f"  ✓  Saved: {out_xlsx.name}  ({out_xlsx.stat().st_size // 1024} KB)")

    # ── Build CIM/RDF ──────────────────────────────────────────────────────────
    CIM  = "http://iec.ch/TC57/CIM100#"
    RDF  = "http://www.w3.org/1999/02/22-rdf-syntax-ns#"
    RDFS = "http://www.w3.org/2000/01/rdf-schema#"
    BASE = f"http://cim-platform.com/subnetwork_{num}/"

    def uid(prefix=""): return prefix + str(uuid.uuid4()).replace("-","")

    lines_rdf = [
        f'<?xml version="1.0" encoding="UTF-8"?>',
        f'<!-- CIM/RDF – Subnetwork {num} – {len(nodes)} nodes, {len(lines)} lines, {len(consumers)} consumers -->',
        f'<rdf:RDF xmlns:rdf="{RDF}" xmlns:rdfs="{RDFS}" xmlns:cim="{CIM}">',
        '',
    ]

    # Base voltages (v_nom is already in kV in PyPSA format)
    bv_map = {}
    for n in nodes:
        v = float(n.get("v_nom", 0.4))  # kV
        key = round(v, 4)
        if key not in bv_map:
            bv_id = uid("BV_")
            bv_map[key] = bv_id
            lines_rdf += [
                f'  <cim:BaseVoltage rdf:ID="{bv_id}">',
                f'    <cim:IdentifiedObject.name>BV_{key}kV</cim:IdentifiedObject.name>',
                f'    <cim:BaseVoltage.nominalVoltage>{key}</cim:BaseVoltage.nominalVoltage>',
                f'  </cim:BaseVoltage>',
            ]

    default_bv_id = list(bv_map.values())[0] if bv_map else None

    # Substation + VoltageLevel
    sub_id = uid("SS_")
    vl_id  = uid("VL_")
    lines_rdf += [
        f'  <cim:Substation rdf:ID="{sub_id}">',
        f'    <cim:IdentifiedObject.name>SS_SubNetwork_{num}</cim:IdentifiedObject.name>',
        f'  </cim:Substation>',
        f'  <cim:VoltageLevel rdf:ID="{vl_id}">',
        f'    <cim:IdentifiedObject.name>VL_SubNetwork_{num}_LV</cim:IdentifiedObject.name>',
        f'    <cim:VoltageLevel.Substation rdf:resource="#{sub_id}"/>',
        f'    <cim:VoltageLevel.BaseVoltage rdf:resource="#{default_bv_id}"/>',
        f'  </cim:VoltageLevel>',
    ]

    # ConnectivityNodes (one per bus)
    cn_map = {}  # bus_id → cn_rdf_id
    for n in nodes:
        bus_id = str(n.get("id_") or n.get("id", ""))
        cn_id  = uid("CN_")
        cn_map[bus_id] = cn_id
        v = float(n.get("v_nom", 0.4))
        bv_id_ref = bv_map.get(round(v, 4), default_bv_id)
        lines_rdf += [
            f'  <cim:ConnectivityNode rdf:ID="{cn_id}">',
            f'    <cim:IdentifiedObject.name>{bus_id}</cim:IdentifiedObject.name>',
            f'    <cim:ConnectivityNode.ConnectivityNodeContainer rdf:resource="#{vl_id}"/>',
            f'  </cim:ConnectivityNode>',
        ]
        bb_id = uid("BB_")
        t_id  = uid("T_")
        lines_rdf += [
            f'  <cim:BusbarSection rdf:ID="{bb_id}">',
            f'    <cim:IdentifiedObject.name>BB_{bus_id}</cim:IdentifiedObject.name>',
            f'    <cim:ConductingEquipment.BaseVoltage rdf:resource="#{bv_id_ref}"/>',
            f'  </cim:BusbarSection>',
            f'  <cim:Terminal rdf:ID="{t_id}">',
            f'    <cim:Terminal.ConductingEquipment rdf:resource="#{bb_id}"/>',
            f'    <cim:Terminal.ConnectivityNode rdf:resource="#{cn_id}"/>',
            f'  </cim:Terminal>',
        ]

    # ACLineSegments
    for idx, l in enumerate(lines):
        ln_id  = uid("L_")
        bus0   = str(l.get("bus0", ""))
        bus1   = str(l.get("bus1", ""))
        r_val  = l.get("r", 0.01)
        s_nom  = l.get("s_nom", 0.1)
        length = l.get("length", 0.1)
        lines_rdf += [
            f'  <cim:ACLineSegment rdf:ID="{ln_id}">',
            f'    <cim:IdentifiedObject.name>{l.get("id_") or "L" + str(idx+1)}</cim:IdentifiedObject.name>',
            f'    <cim:Conductor.length>{length}</cim:Conductor.length>',
            f'    <cim:ACLineSegment.r>{r_val}</cim:ACLineSegment.r>',
            f'  </cim:ACLineSegment>',
        ]
        for seq, bus in enumerate([bus0, bus1]):
            cn_ref = cn_map.get(bus)
            t_id   = uid("T_")
            lines_rdf.append(f'  <cim:Terminal rdf:ID="{t_id}">')
            lines_rdf.append(f'    <cim:Terminal.ConductingEquipment rdf:resource="#{ln_id}"/>')
            if cn_ref:
                lines_rdf.append(f'    <cim:Terminal.ConnectivityNode rdf:resource="#{cn_ref}"/>')
            lines_rdf.append(f'    <cim:ACDCTerminal.sequenceNumber>{seq+1}</cim:ACDCTerminal.sequenceNumber>')
            lines_rdf.append(f'  </cim:Terminal>')

    # EnergyConsumers
    for idx, c in enumerate(consumers):
        ec_id  = uid("EC_")
        bus0   = str(c.get("bus0", ""))
        dp     = c.get("demand_power")
        p_kw   = round(float(dp) / 8760.0, 3) if dp else 0.0
        lines_rdf += [
            f'  <cim:EnergyConsumer rdf:ID="{ec_id}">',
            f'    <cim:IdentifiedObject.name>{c.get("id_") or "LOAD" + str(idx+1)}</cim:IdentifiedObject.name>',
            f'    <cim:EnergyConsumer.p>{p_kw}</cim:EnergyConsumer.p>',
            f'    <cim:EnergyConsumer.q>{round(p_kw*0.2, 3)}</cim:EnergyConsumer.q>',
            f'  </cim:EnergyConsumer>',
        ]
        cn_ref = cn_map.get(bus0)
        t_id   = uid("T_")
        lines_rdf.append(f'  <cim:Terminal rdf:ID="{t_id}">')
        lines_rdf.append(f'    <cim:Terminal.ConductingEquipment rdf:resource="#{ec_id}"/>')
        if cn_ref:
            lines_rdf.append(f'    <cim:Terminal.ConnectivityNode rdf:resource="#{cn_ref}"/>')
        lines_rdf.append(f'  </cim:Terminal>')

    # PowerTransformers
    for idx, t in enumerate(trafos):
        pt_id  = uid("PT_")
        bus0   = str(t.get("bus0", ""))
        bus1   = str(t.get("bus1", ""))
        s_nom  = t.get("s_nom", 0.315)
        lines_rdf += [
            f'  <cim:PowerTransformer rdf:ID="{pt_id}">',
            f'    <cim:IdentifiedObject.name>{t.get("id_") or "TR" + str(idx+1)}</cim:IdentifiedObject.name>',
            f'  </cim:PowerTransformer>',
        ]
        for winding, (bus, seq, v_key) in enumerate([(bus0, 1, "v0"), (bus1, 2, "v1")]):
            ptw = uid("PTW_")
            v   = t.get(v_key, 0.4)
            lines_rdf += [
                f'  <cim:PowerTransformerEnd rdf:ID="{ptw}">',
                f'    <cim:PowerTransformerEnd.PowerTransformer rdf:resource="#{pt_id}"/>',
                f'    <cim:TransformerEnd.endNumber>{seq}</cim:TransformerEnd.endNumber>',
                f'    <cim:PowerTransformerEnd.ratedS>{s_nom}</cim:PowerTransformerEnd.ratedS>',
                f'    <cim:PowerTransformerEnd.ratedU>{v}</cim:PowerTransformerEnd.ratedU>',
                f'  </cim:PowerTransformerEnd>',
            ]
            cn_ref = cn_map.get(bus)
            t_id   = uid("T_")
            lines_rdf.append(f'  <cim:Terminal rdf:ID="{t_id}">')
            lines_rdf.append(f'    <cim:Terminal.ConductingEquipment rdf:resource="#{pt_id}"/>')
            if cn_ref:
                lines_rdf.append(f'    <cim:Terminal.ConnectivityNode rdf:resource="#{cn_ref}"/>')
            lines_rdf.append(f'    <cim:ACDCTerminal.sequenceNumber>{seq}</cim:ACDCTerminal.sequenceNumber>')
            lines_rdf.append(f'  </cim:Terminal>')

    lines_rdf.append("</rdf:RDF>")

    out_rdf = OUT_DIR / f"subnetwork_{num}.rdf"
    out_rdf.write_text("\n".join(lines_rdf), encoding="utf-8")
    print(f"  ✓  Saved: {out_rdf.name}  ({out_rdf.stat().st_size // 1024} KB)")

    return len(nodes), len(lines), len(consumers), len(trafos)


# ── Main ───────────────────────────────────────────────────────────────────────
print("\n" + "═"*60)
print("  Prof File Converter → CIM Platform format")
print("═"*60)

totals = {}
for num, folder in SUBNETWORKS.items():
    if folder.exists():
        totals[num] = convert_subnetwork(num, folder)
    else:
        print(f"\n  SKIP: {folder} not found")

print("\n" + "═"*60)
print("  Summary")
print("═"*60)
for num, (n, l, c, t) in totals.items():
    print(f"  Subnetwork {num}: {n} buses, {l} lines, {c} consumers, {t} transformers")
print(f"\n  Output folder: {OUT_DIR}")
print("\n  To import into the platform:")
print("  1. Upload subnetwork_X.rdf  via  Data Import > CIM/RDF")
print("  OR")
print("  2. Upload subnetwork_X_merged.xlsx  via  Data Import > Excel")
print("═"*60 + "\n")
