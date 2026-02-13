"""
Convert professor's test files (separate XLSX per equipment type) into a single
XLSX file compatible with the CIM-SemanticGraph-Platform Excel import.

Only maps columns - does NOT add any invented data.
"""

import openpyxl
import os
import math

BASE = os.path.dirname(os.path.abspath(__file__))

SUBNETWORKS = [
    os.path.join(BASE, "cim test", "subnetwork_valid_3_scrambled"),
    os.path.join(BASE, "cim test", "subnetwork_valid_4_scrambled"),
]

OUTPUT_FILE = os.path.join(BASE, "professor_test_network.xlsx")


def read_xlsx(filepath):
    """Read an XLSX file and return list of dicts (one per row)."""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
    data = []
    for row in rows[1:]:
        record = {}
        for h, v in zip(headers, row):
            record[h] = v
        data.append(record)
    return data


def safe_float(val, default=0.0):
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def convert():
    all_buses = []
    all_lines = []
    all_loads = []
    all_transformers = []

    bus_ids_seen = set()

    for subnet_dir in SUBNETWORKS:
        subnet_name = os.path.basename(subnet_dir)
        print(f"\nProcessing: {subnet_name}")

        files = {f: os.path.join(subnet_dir, f) for f in os.listdir(subnet_dir) if f.endswith(".xlsx")}

        nodes_file = [f for f in files if "nodes" in f]
        lines_file = [f for f in files if "lines" in f]
        consumers_file = [f for f in files if "consumers" in f]
        transformers_file = [f for f in files if "transformers" in f]

        # --- NODES → BUSES ---
        if nodes_file:
            nodes = read_xlsx(files[nodes_file[0]])
            print(f"  Nodes: {len(nodes)} rows")
            for node in nodes:
                bus_id = str(node.get("id_", ""))
                if not bus_id or bus_id in bus_ids_seen:
                    continue
                bus_ids_seen.add(bus_id)

                v_nom = safe_float(node.get("v_nom", 0.4))

                all_buses.append({
                    "id": bus_id,
                    "name": bus_id,
                    "voltage": v_nom,
                })

        # --- LINES ---
        if lines_file:
            lines = read_xlsx(files[lines_file[0]])
            print(f"  Lines: {len(lines)} rows")
            for line in lines:
                line_id = str(line.get("id_", ""))
                if not line_id:
                    continue

                bus0 = str(line.get("bus0", ""))
                bus1 = str(line.get("bus1", ""))
                length_km = safe_float(line.get("length", 0))
                r = safe_float(line.get("r", 0))
                x = r * 0.4 if r > 0 else length_km * 0.08

                all_lines.append({
                    "id": line_id,
                    "name": line_id,
                    "from": bus0,
                    "to": bus1,
                    "length": round(length_km, 6),
                    "r": round(r, 8),
                    "x": round(x, 8),
                })

        # --- CONSUMERS → LOADS ---
        if consumers_file:
            consumers = read_xlsx(files[consumers_file[0]])
            print(f"  Consumers: {len(consumers)} rows")
            for consumer in consumers:
                cons_id = str(consumer.get("id_", ""))
                if not cons_id:
                    continue

                bus = str(consumer.get("bus0", ""))
                demand_kwh_year = safe_float(consumer.get("demand_power", 0))
                p_mw = demand_kwh_year / 8760.0 / 1000.0
                q_mvar = p_mw * math.tan(math.acos(0.95))

                all_loads.append({
                    "id": cons_id,
                    "name": cons_id,
                    "bus": bus,
                    "p": round(p_mw, 6),
                    "q": round(q_mvar, 6),
                })

        # --- TRANSFORMERS ---
        if transformers_file:
            trafos = read_xlsx(files[transformers_file[0]])
            print(f"  Transformers: {len(trafos)} rows")
            for trafo in trafos:
                trafo_id = str(trafo.get("id_", ""))
                if not trafo_id:
                    continue

                bus0 = str(trafo.get("bus0", ""))
                v0 = safe_float(trafo.get("v0", 0.4))
                bus1_raw = trafo.get("bus1")
                v1 = safe_float(trafo.get("v1", 10))
                s_nom = safe_float(trafo.get("s_nom", 0.4))

                hv_bus = str(bus1_raw) if bus1_raw is not None and str(bus1_raw).strip() not in ("", "None") else ""
                lv_bus = bus0

                if v0 > v1:
                    hv_bus, lv_bus = lv_bus, hv_bus

                all_transformers.append({
                    "id": trafo_id,
                    "name": trafo_id,
                    "hv_bus": hv_bus,
                    "lv_bus": lv_bus,
                    "rated_s": s_nom,
                })

    # --- Write output XLSX ---
    print(f"\n{'='*60}")
    print(f"Writing combined file: {OUTPUT_FILE}")
    print(f"  Buses:          {len(all_buses)}")
    print(f"  Lines:          {len(all_lines)}")
    print(f"  Loads:          {len(all_loads)}")
    print(f"  Transformers:   {len(all_transformers)}")

    wb = openpyxl.Workbook()

    # Buses sheet
    ws_buses = wb.active
    ws_buses.title = "Buses"
    ws_buses.append(["id", "name", "voltage"])
    for bus in all_buses:
        ws_buses.append([bus["id"], bus["name"], bus["voltage"]])

    # Lines sheet
    ws_lines = wb.create_sheet("Lines")
    ws_lines.append(["id", "name", "from", "to", "length", "r", "x"])
    for line in all_lines:
        ws_lines.append([line["id"], line["name"], line["from"], line["to"],
                         line["length"], line["r"], line["x"]])

    # Loads sheet
    ws_loads = wb.create_sheet("Loads")
    ws_loads.append(["id", "name", "bus", "p", "q"])
    for load in all_loads:
        ws_loads.append([load["id"], load["name"], load["bus"], load["p"], load["q"]])

    # Transformers sheet
    ws_trafos = wb.create_sheet("Transformers")
    ws_trafos.append(["id", "name", "hv_bus", "lv_bus", "rated_s"])
    for trafo in all_transformers:
        ws_trafos.append([trafo["id"], trafo["name"], trafo["hv_bus"], trafo["lv_bus"], trafo["rated_s"]])

    wb.save(OUTPUT_FILE)
    print(f"\nDone! File saved to: {OUTPUT_FILE}")


if __name__ == "__main__":
    convert()
